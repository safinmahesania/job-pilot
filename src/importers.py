"""Bringing jobs in from outside the fetch pipeline.

Three routes, one destination:

  * `import_tabular`  — a CSV or Excel file (any column names; they are matched
    by alias).
  * `import_text`     — you paste a whole job post; the model pulls the fields out.
  * `parse_email_file` — you hand JobPilot a job-alert email you exported
    (Save as .eml or .html), or drop one into data/mail_drop/.

Why email at all: LinkedIn and Indeed no longer have a usable public job API,
both block automated search, and scraping LinkedIn is against its terms. But both
will happily *email you* the results of a saved search. That email is yours.

Why JobPilot does NOT read your mailbox
---------------------------------------
An earlier version connected over IMAP. It was removed on purpose. IMAP has no
folder-scoped credential: an app password can read the entire account, so even
code that only ever opens one folder is one bug — or one leaked .env — away from
exposing everything. That is not a risk worth taking for a convenience.

So JobPilot has no mail credentials, no IMAP client, and no network path to your
mail. You export the alert emails you want it to see and hand them over. It reads
exactly what you give it and nothing else. This is slightly less convenient and
categorically safer.

The description problem, and what we do about it
------------------------------------------------
Alert emails carry a title, a company, a location and a link — no job
description. Our scoring is an AI judgement of fit, and fit cannot be judged
from a title alone; a score derived from nothing would look authoritative and be
worthless.

So: if a real description can be recovered, the job is scored normally. If it
cannot, the job is stored **unscored** and surfaced in its own tab for you to
triage by hand. It is never given a made-up number.

Recovery works because most LinkedIn/Indeed listings are really hosted on an ATS
(Greenhouse, Lever, Ashby…). Following the link's redirects usually lands on that
ATS page, which we can read. LinkedIn's and Indeed's own pages are behind a login
or bot wall — those stay unscored, by design.

Mail privacy
------------
Inbox import is deliberately narrow. It reads exactly one folder — the one you
name in IMAP_FOLDER — and nothing else. Point it at a dedicated folder (a Gmail
label with a filter that routes job alerts into it) and JobPilot never has sight
of the rest of your mail, not even your inbox.

Within that folder it is read-only: the mailbox is opened with readonly=True, so
nothing is deleted, moved, marked read, or written back. Only the fields below
are extracted (title, company, location, link) and stored locally in your SQLite
database. Nothing is sent anywhere. Message bodies are not kept.

By default it further filters to known job-alert senders, so even a shared folder
yields only alert mail — but the folder boundary is the real guarantee.
"""
import csv
import io
import re

import httpx
from bs4 import BeautifulSoup

from src import llm, store
from src.logs import log
from src.config import load_profile
from src.normalize import normalize, clean_html
from src.scoring.prefilter import passes
from src.scoring.rerank import score_job, build_calibration
from src.paths import DEFAULT_SCORE_THRESHOLD


# ── Column aliases for tabular imports ──────────────────────────────────────
# Spreadsheets come from everywhere and name things differently. Match on a
# normalised header, so "Job Title", "job_title" and "position" all land in the
# same field.

ALIASES = {
    "title": ["title", "job title", "job", "position", "role", "job name"],
    "company": ["company", "employer", "organisation", "organization", "company name"],
    "location": ["location", "city", "place", "where", "job location"],
    "apply_url": ["apply url", "apply link", "url", "link", "job url", "job link",
                  "posting url", "application link", "href"],
    "description": ["description", "job description", "details", "summary",
                    "jd", "content", "body", "text"],
    "posted_date": ["posted", "posted date", "date", "date posted", "published"],
    "job_type": ["job type", "type", "employment type", "contract"],
    "salary": ["salary", "pay", "compensation", "salary range"],
    "source": ["source", "board", "site", "via"],
}

_ALIAS_LOOKUP = {alias: field for field, aliases in ALIASES.items() for alias in aliases}


def _norm_header(h: str) -> str:
    return re.sub(r"[^a-z0-9 ]", " ", str(h or "").lower()).strip()


def _map_row(row: dict) -> dict:
    """Turn one spreadsheet row into our field names."""
    out = {}
    for header, value in row.items():
        field = _ALIAS_LOOKUP.get(_norm_header(header))
        if field and value not in (None, ""):
            out[field] = str(value).strip()
    return out


# ── Route 1: CSV / Excel ────────────────────────────────────────────────────

def parse_tabular(data: bytes, filename: str) -> list[dict]:
    """Read a CSV or Excel file into mapped job dicts."""
    name = (filename or "").lower()

    if name.endswith((".xlsx", ".xlsm", ".xls")):
        try:
            from openpyxl import load_workbook
        except ImportError as e:
            raise RuntimeError(
                "Excel import needs openpyxl. Install it with: pip install openpyxl"
            ) from e
        wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
        sheet = wb.active
        rows = sheet.iter_rows(values_only=True)
        headers = [str(h or "") for h in next(rows, [])]
        records = [dict(zip(headers, r)) for r in rows]
    else:
        text = data.decode("utf-8-sig", errors="replace")
        records = list(csv.DictReader(io.StringIO(text)))

    mapped = [_map_row(r) for r in records]
    return [m for m in mapped if m.get("title") and m.get("company")]


# ── Route 2: pasted job post ────────────────────────────────────────────────

_PARSE_SYSTEM = """You extract structured data from a pasted job posting.

Return ONLY a JSON object with these keys:
  title, company, location, apply_url, job_type, salary, posted_date, description

Rules:
- Copy facts from the text. Never invent a company, a location or a URL.
- If something is not in the text, use an empty string. Do not guess.
- `description` is the body of the posting — responsibilities, requirements,
  qualifications. Keep it substantially complete; do not summarise it away, and
  do not add anything that isn't there. Strip navigation, cookie banners, "apply
  now" boilerplate and share buttons.
- `job_type` is one of: Full-time, Part-time, Contract, Internship, or "" if unclear.

No prose, no markdown — the JSON object only."""


def parse_text(text: str) -> dict:
    """Pull a job out of a pasted posting. Grounded: blanks, not guesses."""
    import json

    if not text or len(text.strip()) < 50:
        raise ValueError("that's too short to be a job posting")

    body, _ = llm.generate(_PARSE_SYSTEM, text.strip()[:20000])
    match = re.search(r"\{.*\}", body, re.S)
    if not match:
        raise ValueError("couldn't read a job out of that text")

    data = json.loads(match.group(0))
    if not data.get("title") or not data.get("company"):
        raise ValueError("couldn't find a job title and company in that text")

    return {
        "source": "pasted",
        "title": data.get("title", "").strip(),
        "company": data.get("company", "").strip(),
        "location": data.get("location", "").strip() or "Not specified",
        "apply_url": data.get("apply_url", "").strip(),
        "description": data.get("description", "").strip(),
        "job_type": data.get("job_type", "").strip() or "Unknown",
        "posted_date": data.get("posted_date", "").strip(),
    }


# ── Route 3: job-alert emails ───────────────────────────────────────────────

# Only mail from these senders is treated as a job alert.
ALERT_SENDERS = (
    "linkedin.com",
    "indeed.com",
    "indeedemail.com",
    "glassdoor.com",
    "ziprecruiter.com",
)


def parse_email_file(data: bytes, filename: str = "") -> list[dict]:
    """Read a job-alert email you exported and pull the job cards out of it.

    Accepts .eml (Save as / Download message) and .html (Save page). Nothing is
    fetched, nothing is logged in, nothing else is read.
    """
    import email as email_mod
    from email.header import decode_header

    name = (filename or "").lower()

    if name.endswith(".eml") or data[:5] in (b"From ", b"Retur", b"Deliv", b"MIME-"):
        msg = email_mod.message_from_bytes(data)
        sender = (msg.get("From") or "").lower()
        subject = ""
        try:
            part, enc = decode_header(msg.get("Subject", ""))[0]
            subject = (part.decode(enc or "utf-8", "replace")
                       if isinstance(part, bytes) else part)
        except Exception:
            pass
        html = _email_html(msg)
    else:
        html = data.decode("utf-8", errors="replace")
        sender = ""
        subject = filename

    if not html:
        return []

    # Two shapes of alert email, and one inbox usually gets both: aggregators (LinkedIn,
    # Indeed…) link to their own walled listing, while employers (Deloitte, Celestica,
    # Scotiabank…) link to a /job/ path on their own careers domain. Run both parsers and
    # keep whatever each finds — an email is only ever one shape, so they don't overlap,
    # and dedupe by URL cleans up any that do.
    jobs = _parse_alert_html(html, sender, subject)
    jobs += _parse_career_site_html(html, sender, subject)
    return _dedupe_by_url(jobs)


def read_mail_drop() -> tuple[list[dict], list[str]]:
    """Ingest every email sitting in data/mail_drop/.

    Drag exported alert emails into that folder and run this. Files are read and
    left where they are — JobPilot does not delete your mail.
    """
    from src.paths import MAIL_DROP_DIR

    MAIL_DROP_DIR.mkdir(parents=True, exist_ok=True)
    jobs, files = [], []

    for path in sorted(MAIL_DROP_DIR.iterdir()):
        if not path.is_file():
            continue
        if path.suffix.lower() not in (".eml", ".html", ".htm", ".txt"):
            continue
        try:
            jobs += parse_email_file(path.read_bytes(), path.name)
            files.append(path.name)
        except Exception as e:
            log.warning("[mail_drop] %s: %s", path.name, e)

    return _dedupe_by_url(jobs), files


def _email_html(msg) -> str:
    """The HTML body of an email, whatever shape it arrived in."""
    if msg.is_multipart():
        for part in msg.walk():
            if part.get_content_type() == "text/html":
                payload = part.get_payload(decode=True)
                if payload:
                    return payload.decode(part.get_content_charset() or "utf-8", "replace")
        return ""
    if msg.get_content_type() == "text/html":
        payload = msg.get_payload(decode=True)
        if payload:
            return payload.decode(msg.get_content_charset() or "utf-8", "replace")
    return ""


# Links in these emails are tracking redirects; the job ones share a shape.
_JOB_LINK = re.compile(r"(linkedin\.com/(comm/)?jobs/view|indeed\.com/(rc/clk|viewjob|pagead)"
                       r"|glassdoor\.[a-z.]+/(job|partner)|ziprecruiter\.com/(jobs|c/))", re.I)

# Company career sites (Deloitte, Celestica, Scotiabank, and every other site on the
# jobs2web / SuccessFactors / Workday family) put the posting under a /job/ path on
# their own domain: careers.example.com/job/Some-Title/12345/. Matching the path rather
# than a list of companies means one rule covers every employer who sends these alerts —
# and there are far more of them than there are aggregators.
_CAREER_JOB_LINK = re.compile(r"/job(?:s)?/[^/?#]+", re.I)

# Aggregators are handled by _JOB_LINK above and are behind login walls; never treat one
# as a company career site.
_WALLED_HOSTS = ("linkedin.com", "indeed.com", "glassdoor.", "ziprecruiter.com",
                 "monster.", "simplyhired.")

# Paths on a careers domain that are not postings.
_NOT_A_JOB = re.compile(r"/(unsubscribe|social-matcher|privacy|search|login|alerts?)\b", re.I)

# A location often trails the title on these sites: "Senior AI Engineer - Toronto, ON, CA".
_LOCATION_TAIL = re.compile(
    r"\s[-–]\s([A-Za-z .'/]+,\s*[A-Z]{2}(?:,\s*[A-Z]{2})?(?:,\s*[A-Z0-9]+)?)\s*$")


def _company_from_host(url: str) -> str:
    """Derive the employer from a careers domain: careers.deloitte.ca -> Deloitte."""
    m = re.match(r"https?://([^/]+)", url or "")
    if not m:
        return ""
    host = m.group(1).lower()
    host = re.sub(r"^(careers?|jobs|recruiting|talent)\.", "", host)
    host = re.sub(r"\.(com|ca|co\.uk|org|net|eu|io|jobs)$", "", host)
    name = host.split(".")[0].replace("-", " ").strip()
    return name.title() if name else ""


def _parse_career_site_html(html: str, sender: str, subject: str) -> list[dict]:
    """Pull postings out of an employer's own job-alert email.

    These alerts (Deloitte, Celestica, Scotiabank, …) all come from the same family of
    notification platforms and share one structure: each posting is an <a> whose href is
    a /job/ path on the company's careers domain, with the title as the link text and the
    location frequently appended to it. The employer is the domain itself.

    The aggregator parser can't see these at all — its pattern only knows LinkedIn,
    Indeed, Glassdoor and ZipRecruiter — so without this every company alert imported
    zero jobs.
    """
    soup = BeautifulSoup(html, "html.parser")
    out = []

    for anchor in soup.find_all("a", href=True):
        href = anchor["href"]
        low = href.lower()
        if any(w in low for w in _WALLED_HOSTS):
            continue
        if not _CAREER_JOB_LINK.search(href) or _NOT_A_JOB.search(href):
            continue

        text = anchor.get_text(" ", strip=True)
        if not text or len(text) < 5 or len(text) > 160:
            continue
        if re.match(r"^(view|apply|see all|show more|unsubscribe|click here)", text, re.I):
            continue

        # Split "Job Title - Toronto, ON, CA" into its two halves when the tail really
        # looks like a location; plenty of titles contain a dash and no location at all
        # ("Software Engineer - Platform Engineering"), so the pattern has to be strict.
        m = _LOCATION_TAIL.search(text)
        title = (text[:m.start()] if m else text).strip()
        location = m.group(1).strip() if m else ""

        company = _company_from_host(href)
        out.append({
            "source": company.lower() or (sender.split(".")[0] if sender else "careers"),
            "title": title,
            "company": company or "(unknown)",
            "location": location or "Not specified",
            "apply_url": href,
            "description": "",            # recovered later from the posting itself
            "job_type": "Unknown",
        })
    return out


def _parse_alert_html(html: str, sender: str, subject: str) -> list[dict]:
    """Pull job cards out of an alert email.

    Deliberately structure-agnostic: rather than chase each provider's table
    layout (which they change often), find the links that point at a job and read
    the text around them. That survives redesigns.
    """
    soup = BeautifulSoup(html, "html.parser")
    source = "linkedin" if "linkedin" in sender else \
             "indeed" if "indeed" in sender else \
             sender.split(".")[0]

    out = []
    for anchor in soup.find_all("a", href=True):
        href = anchor["href"]
        if not _JOB_LINK.search(href):
            continue

        title = anchor.get_text(" ", strip=True)
        if not title or len(title) < 3 or len(title) > 120:
            continue
        if re.match(r"^(view|apply|see all|show more|unsubscribe)", title, re.I):
            continue

        # Company and location live in the text just after the title link.
        # Company and location sit together on their own line, separated by a middle
        # dot: "SecureRx Technologies Inc. · Ontario, Canada (Remote)". Finding that line
        # is far steadier than taking the first two lines by position — when the anchor
        # text and the title line differ, position-based reading put the *title* in the
        # company field, which is where entries like "Data Engineering Intern" as a
        # company name came from.
        company, location = "", ""
        block = anchor.find_parent(["td", "div", "table"])
        if block:
            lines = [ln.strip() for ln in block.get_text("\n", strip=True).split("\n")
                     if ln.strip() and ln.strip() != title]
            for ln in lines:
                if "·" in ln:
                    left, _, right = ln.partition("·")
                    company, location = left.strip()[:100], right.strip()[:100]
                    break
            if not company and lines:            # no dot -> fall back to the old reading
                company = lines[0][:100]
                if len(lines) > 1:
                    location = lines[1][:100]

        # The anchor text sometimes runs the whole card together — "Product Intern Sezzle
        # · Canada (Remote) Easy Apply" — so cut the title back to what precedes the
        # company, and drop the badges LinkedIn appends after it.
        if company and company in title:
            title = title.split(company)[0].strip(" -–·,")
        title = re.sub(r"\s*(easy apply|be an early applicant|promoted|\d+ school alums?)\s*$",
                       "", title, flags=re.I).strip()
        if not title:
            continue

        out.append({
            "source": source,
            "title": title,
            "company": company or "(unknown)",
            "location": location or "Not specified",
            "apply_url": href,
            "description": "",            # alerts carry none — recovered later
            "job_type": "Unknown",
        })
    return out


def _dedupe_by_url(jobs: list[dict]) -> list[dict]:
    seen, out = set(), []
    for j in jobs:
        key = (j.get("apply_url") or "")[:200]
        if key and key not in seen:
            seen.add(key)
            out.append(j)
    return out


# ── Recovering a description so a job can be scored ─────────────────────────

# Hosts we can actually read. LinkedIn and Indeed serve a login/bot wall to
# anything automated, so we don't pretend otherwise — those stay unscored.
READABLE_HOSTS = (
    "greenhouse.io", "lever.co", "ashbyhq.com", "myworkdayjobs.com",
    "oraclecloud.com", "smartrecruiters.com", "workable.com", "bamboohr.com",
    "jobvite.com", "icims.com", "recruitee.com", "teamtailor.com",
    "successfactors.com", "taleo.net", "breezy.hr",
)


def follow_links_enabled() -> bool:
    """Whether tracking redirects from alert emails may be followed."""
    from src.paths import FOLLOW_JOB_LINKS
    try:
        conn = store.connect()
        val = store.get_setting(conn, "follow_job_links", None)
        conn.close()
        if val is not None:
            return val == "1"
    except Exception:
        pass
    return FOLLOW_JOB_LINKS


def _is_readable(url: str) -> bool:
    """Whether a posting at this URL may be read.

    Two kinds of host qualify. The named ATS platforms below are the classic case. So is
    an employer's own careers domain — careers.deloitte.ca, jobs.scotiabank.com — which
    publishes its descriptions openly; that IS the site's purpose. Excluding them was
    why every company job-alert import stayed unscored.

    Aggregators are the exception in both directions: LinkedIn, Indeed and their peers
    serve a login or bot wall to anything automated, so nothing is ever read from them.
    """
    low = (url or "").lower()
    if any(w in low for w in _WALLED_HOSTS):
        return False
    if any(host in low for host in READABLE_HOSTS):
        return True
    # An employer's careers site: a /job/ path on a careers-ish host.
    m = re.match(r"https?://([^/]+)", low)
    host = m.group(1) if m else ""
    looks_like_careers = bool(re.match(r"^(careers?|jobs|recruiting|talent)\.", host))
    return looks_like_careers and bool(_CAREER_JOB_LINK.search(low))


def recover_description(url: str, timeout: int = 15) -> str:
    """Read the posting behind a link, if the host publishes it openly.

    Alert-email links are tracking redirects. Following one tells LinkedIn or
    Indeed that this message was engaged with, and gives them the request's IP.

    The honest reckoning is that this costs you almost nothing: you were going to
    click that link yourself to read the job, from this same machine, and the
    tracker would fire then. The one real difference is that the request happens
    even for jobs you never look at. Setting FOLLOW_JOB_LINKS = False (or the
    Settings toggle) stops it — at the price of every alert-email job staying
    unscored, because a description is what scoring needs.

    Either way, the response is only read when the redirect lands on a host that
    publishes job descriptions openly. LinkedIn's and Indeed's own pages are
    behind a wall; nothing is scraped from them.

    Returns "" when the description can't be had. That is a normal outcome, not an
    error: an unscored job is honest, a job scored from its title is not.
    """
    if not url:
        return ""

    direct = _is_readable(url)

    # A tracking link needs the redirect followed to reach the real posting.
    if not direct and not follow_links_enabled():
        return ""

    try:
        r = httpx.get(url, follow_redirects=True, timeout=timeout,
                      headers={"User-Agent": "Mozilla/5.0 (JobPilot)"})

        # Whatever the redirect chain, only read a host that publishes openly.
        if not _is_readable(str(r.url)):
            return ""
        if r.status_code != 200:
            return ""

        soup = BeautifulSoup(r.text, "html.parser")
        for tag in soup(["script", "style", "nav", "header", "footer", "svg"]):
            tag.decompose()

        # Prefer the posting's own container; fall back to the main body.
        node = (soup.select_one("#content, .content, [class*='job-description'], "
                                "[class*='jobDescription'], [data-automation-id='jobPostingDescription'], "
                                "main, article") or soup.body)
        if not node:
            return ""

        html = str(node)
        text = BeautifulSoup(html, "html.parser").get_text(" ", strip=True)
        if len(text) < 400:                 # a stub or a wall, not a description
            return ""
        return clean_html(html)[:30000]
    except Exception:
        return ""


# ── The shared import path ──────────────────────────────────────────────────

def import_jobs(raw_jobs: list[dict], *, source: str = "import",
                fetch_missing: bool = True) -> dict:
    """Normalise, dedupe, score where possible, and store.

    A job is scored only when it has a real description. Otherwise it is stored
    unscored (score = NULL) and appears in the Unscored tab for manual triage.
    """
    profile = load_profile()
    conn = store.connect()
    threshold = int(store.get_setting(conn, "score_threshold", DEFAULT_SCORE_THRESHOLD))
    scoring_on = store.get_setting(conn, "scoring_enabled", "1") == "1"

    # Imported jobs are scored against the same calibration as fetched ones.
    calibration = build_calibration() if scoring_on else ""

    stats = {"seen": 0, "imported": 0, "scored": 0, "unscored": 0,
             "duplicates": 0, "errors": 0, "dropped": 0}

    for raw in raw_jobs:
        stats["seen"] += 1
        raw.setdefault("source", source)

        try:
            job = normalize(raw)

            if store.already_seen(conn, job["dedupe_hash"]):
                stats["duplicates"] += 1
                continue

            # No description? Try to recover one from the link before giving up.
            if fetch_missing and not job.get("description"):
                recovered = recover_description(job.get("apply_url") or "")
                if recovered:
                    job["description"] = recovered

            # Location gate — runs before scoring, and works without a description
            # because alert emails carry a location. A job we can SEE is outside Canada
            # and not remote is dropped now rather than parked in the unscored tab.
            #
            # It only fires on a location we actually have. Some employers' alerts
            # (Deloitte's, for one) name no location at all, and those postings are on a
            # Canadian careers site and perfectly relevant — dropping them for a missing
            # field would silently throw away good jobs. Unknown location therefore means
            # "keep and let scoring judge it", not "discard".
            from src.scoring.prefilter import _check_locations
            allowed_locations = (profile.get("constraints") or {}).get("locations")
            loc_known = (job.get("location") or "").strip().lower() not in (
                "", "not specified", "unknown", "n/a", "-")
            if (allowed_locations and loc_known
                    and not _check_locations(job, allowed_locations)):
                store.mark_seen(conn, job["dedupe_hash"], "trashed")
                stats["dropped"] = stats.get("dropped", 0) + 1
                continue

            has_jd = bool((job.get("description") or "").strip())

            if has_jd and scoring_on and passes(job, profile):
                result = score_job(job, profile, calibration)
                if result is not None:
                    job.update(score=result.overall,
                               skills_score=result.skills_score,
                               seniority_score=result.seniority_score,
                               domain_score=result.domain_score,
                               rationale=result.rationale, flags=None)
                    store.save_job(conn, job)
                    store.mark_seen(conn, job["dedupe_hash"],
                                    "kept" if result.overall >= threshold else "trashed",
                                    result.overall)
                    stats["imported"] += 1
                    stats["scored"] += 1
                    continue

            # Everything else lands unscored: no description, scoring off, or the
            # model failed. The job is still yours to look at — it just doesn't
            # pretend to have been judged.
            job.update(score=None, skills_score=None, seniority_score=None,
                       domain_score=None, rationale=None, flags=None)
            store.save_job(conn, job)
            store.mark_seen(conn, job["dedupe_hash"], "kept")
            stats["imported"] += 1
            stats["unscored"] += 1

        except Exception as e:
            log.warning("[import] %s: %s", raw.get("title", "?"), e)
            stats["errors"] += 1

    conn.commit()
    conn.close()
    return stats
